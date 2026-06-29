"""
LeadRadar — Main API v1
Routes:
  /api/v1/auth/*        — register, login, forgot/reset password
  /api/v1/categories/*  — dynamic categories from DB
  /api/v1/search        — Overpass API city scan (protected)
  /api/v1/export        — Excel export (protected)
  /api/v1/leads/*       — save/fetch/delete scan results (protected)
"""

from fastapi import FastAPI, HTTPException, Query, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from auth       import router as authRouter, verifyToken
from categories import router as categoriesRouter
from leads      import router as leadsRouter
from features   import router as featuresRouter
from howItWorks   import router as howItWorksRouter
from faqs           import router as faqsRouter
from database   import executeQuery
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="LeadRadar API", version="1.0.0", redirect_slashes=False)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Mount routers ──────────────────────────────────────────────────────────────
app.include_router(authRouter,       prefix="/api/v1/auth",       tags=["Auth"])
app.include_router(categoriesRouter, prefix="/api/v1/categories", tags=["Categories"])
app.include_router(leadsRouter,      prefix="/api/v1/leads",      tags=["Leads"])
app.include_router(featuresRouter,   prefix="/api/v1/features",   tags=["Features"])
app.include_router(howItWorksRouter, prefix="/api/v1/how-it-works", tags=["HowItWorks"])
app.include_router(faqsRouter,       prefix="/api/v1/faqs",        tags=["FAQs"])

OVERPASS_URL  = "https://overpass-api.de/api/interpreter"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"


# ── Load category map from DB for Overpass queries ────────────────────────────

def getCategoryMap() -> dict:
    """Fetch all categories from DB and build a key→tags map."""
    rows = executeQuery(
        "SELECT categoryKey, label, amenityTags, shopTags FROM categories WHERE isActive = 1"
    )
    return {
        row["categoryKey"]: {
            "amenity": row["amenityTags"],
            "shop":    row["shopTags"],
            "label":   row["label"],
        }
        for row in rows
    }


# ── Overpass query builders ────────────────────────────────────────────────────

def buildFilters(amenityTags, shopTags) -> str:
    lines = []
    for tagType, tags in [("amenity", amenityTags), ("shop", shopTags)]:
        if tags:
            lines.append(f'node["{tagType}"~"{tags}"](area.searchArea);')
            lines.append(f'way["{tagType}"~"{tags}"](area.searchArea);')
    return "\n      ".join(lines)


def buildQueryByAreaName(city, amenityTags, shopTags) -> str:
    filters = buildFilters(amenityTags, shopTags)
    return f"""
    [out:json][timeout:60];
    area[name="{city}"]["admin_level"~"4|5|6|7|8"]->.searchArea;
    (
      {filters}
    );
    out body;
    >;
    out skel qt;
    """


def buildQueryByRelationId(relationId, amenityTags, shopTags) -> str:
    areaId  = relationId + 3_600_000_000
    filters = buildFilters(amenityTags, shopTags)
    return f"""
    [out:json][timeout:60];
    area({areaId})->.searchArea;
    (
      {filters}
    );
    out body;
    >;
    out skel qt;
    """


def buildQueryByBbox(south, west, north, east, amenityTags, shopTags) -> str:
    lines = []
    for tagType, tags in [("amenity", amenityTags), ("shop", shopTags)]:
        if tags:
            lines.append(f'node["{tagType}"~"{tags}"]({south},{west},{north},{east});')
            lines.append(f'way["{tagType}"~"{tags}"]({south},{west},{north},{east});')
    filterStr = "\n      ".join(lines)
    return f"""
    [out:json][timeout:60];
    (
      {filterStr}
    );
    out body;
    >;
    out skel qt;
    """


def resolveCity(city: str) -> dict:
    try:
        resp = requests.get(
            NOMINATIM_URL,
            params={"q": city, "format": "json", "limit": 5, "addressdetails": 1},
            headers={"User-Agent": "LeadRadar/2.0"},
            timeout=10,
        )
        results   = resp.json()
        preferred = None
        for r in results:
            if r.get("type") in ("city", "town", "municipality", "administrative"):
                preferred = r
                break
        if not preferred and results:
            preferred = results[0]
        if not preferred:
            return {}

        relationId = int(preferred["osm_id"]) if preferred.get("osm_type") == "relation" else None
        bbox = preferred.get("boundingbox")
        if bbox:
            south, north, west, east = float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])
            return {
                "relationId": relationId,
                "bbox": (south - 0.01, west - 0.01, north + 0.01, east + 0.01),
            }
    except Exception as e:
        print(f"Nominatim error: {e}")
    return {}


def runOverpassQuery(query: str) -> list:
    resp = requests.post(
        OVERPASS_URL,
        data={"data": query},
        timeout=65,
        headers={"User-Agent": "LeadRadar/2.0"},
    )
    resp.raise_for_status()
    return resp.json().get("elements", [])


def fetchPlaces(city: str, category: str = "all") -> list:
    """3-layer fallback: area name → relation ID → bounding box."""
    categoryMap = getCategoryMap()
    cat         = categoryMap.get(category, categoryMap.get("all", {}))
    amenityTags = cat.get("amenity")
    shopTags    = cat.get("shop")
    elements    = []

    try:
        elements = runOverpassQuery(buildQueryByAreaName(city, amenityTags, shopTags))
        print(f"[Layer 1] {len(elements)} elements")
    except Exception as e:
        print(f"[Layer 1] Failed: {e}")

    if not elements:
        cityInfo = resolveCity(city)

        if cityInfo.get("relationId"):
            try:
                elements = runOverpassQuery(
                    buildQueryByRelationId(cityInfo["relationId"], amenityTags, shopTags)
                )
                print(f"[Layer 2] {len(elements)} elements")
            except Exception as e:
                print(f"[Layer 2] Failed: {e}")

        if not elements and cityInfo.get("bbox"):
            try:
                s, w, n, e = cityInfo["bbox"]
                elements = runOverpassQuery(buildQueryByBbox(s, w, n, e, amenityTags, shopTags))
                print(f"[Layer 3] {len(elements)} elements")
            except Exception as e:
                print(f"[Layer 3] Failed: {e}")

    if not elements:
        raise HTTPException(
            status_code=404,
            detail=f"No results for '{city}'. Try a different city or category."
        )

    seen   = set()
    places = []
    for el in elements:
        if el.get("type") not in ("node", "way"):
            continue
        parsed = parsePlace(el)
        if not parsed:
            continue
        key = f"{parsed['name'].lower()}_{parsed['type']}"
        if key in seen:
            continue
        seen.add(key)
        places.append(parsed)

    places.sort(key=lambda p: (p["hasWebsite"], p["name"].lower()))
    return places


def parsePlace(element: dict) -> dict | None:
    tags    = element.get("tags", {})
    name    = tags.get("name", "").strip()
    if not name:
        return None

    amenity = tags.get("amenity", "")
    shop    = tags.get("shop", "")
    typeStr = (amenity or shop).replace("_", " ").title()
    phone   = tags.get("phone", tags.get("contact:phone", ""))
    website = tags.get("website", tags.get("contact:website", ""))
    hours   = tags.get("opening_hours", "")
    address = buildAddress(tags)

    lat = element.get("lat") or element.get("center", {}).get("lat")
    lon = element.get("lon") or element.get("center", {}).get("lon")
    mapsUrl = (
        f"https://www.google.com/maps?q={lat},{lon}" if lat and lon
        else f"https://www.google.com/maps/search/{requests.utils.quote(name)}"
    )

    return {
        "id":         element.get("id"),
        "name":       name,
        "type":       typeStr,
        "address":    address,
        "phone":      phone,
        "website":    website,
        "hours":      hours,
        "hasWebsite": bool(website),
        "mapsUrl":    mapsUrl,
        "lat":        lat,
        "lon":        lon,
    }


def buildAddress(tags: dict) -> str:
    parts = []
    for key in ["addr:housenumber", "addr:street", "addr:suburb", "addr:city", "addr:state"]:
        val = tags.get(key, "").strip()
        if val:
            parts.append(val)
    return ", ".join(parts) if parts else tags.get("addr:full", "")


# ── Search & Export routes ─────────────────────────────────────────────────────

@app.get("/api/v1/search")
def searchPlaces(
    city:        str  = Query(..., min_length=2),
    category:    str  = Query("all"),
    currentUser: dict = Depends(verifyToken),
):
    """
    GET /api/v1/search?city=Jodhpur&category=restaurant
    Protected. Returns live Overpass results — not from DB.
    User can then choose to save via POST /api/v1/leads/save.
    """
    categoryMap = getCategoryMap()
    cat         = categoryMap.get(category, categoryMap.get("all", {}))

    places     = fetchPlaces(city, category)
    noWebsite  = [p for p in places if not p["hasWebsite"]]
    hasWebsite = [p for p in places if p["hasWebsite"]]

    return {
        "city":          city,
        "category":      category,
        "categoryLabel": cat.get("label", category),
        "total":         len(places),
        "noWebsite":     len(noWebsite),
        "hasWebsite":    len(hasWebsite),
        "places":        places,
        "generatedAt":   datetime.utcnow().isoformat(),
    }


@app.get("/api/v1/export")
def exportExcel(
    city:        str  = Query(..., min_length=2),
    category:    str  = Query("all"),
    currentUser: dict = Depends(verifyToken),
):
    """
    GET /api/v1/export?city=Jodhpur&category=restaurant
    Protected. Downloads Excel of live search results.
    """
    categoryMap = getCategoryMap()
    cat         = categoryMap.get(category, categoryMap.get("all", {}))

    places    = fetchPlaces(city, category)
    noWebsite = [p for p in places if not p["hasWebsite"]]

    wb     = buildExcel(places, noWebsite, city, cat.get("label", category))
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"leadradar_{city.lower().replace(' ', '_')}_{category}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/v1/health")
def health():
    return {"status": "ok", "service": "LeadRadar API v1"}


# ── Excel builder ──────────────────────────────────────────────────────────────

def buildExcel(allPlaces: list, noWebsite: list, city: str, catLabel: str):
    wb         = openpyxl.Workbook()
    ws1        = wb.active
    ws1.title  = "Pitch Targets"

    darkFill   = PatternFill("solid", start_color="262626")
    accentFill = PatternFill("solid", start_color="AABBC5")
    altFill    = PatternFill("solid", start_color="F2F4F5")
    thinBorder = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    ws1.merge_cells("A1:H1")
    t           = ws1["A1"]
    t.value     = f"LeadRadar — {catLabel} in {city}  |  {len(noWebsite)} pitch targets"
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    t.fill      = darkFill
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:H2")
    s           = ws1["A2"]
    s.value     = f"Generated by LeadRadar · {datetime.now().strftime('%d %B %Y')} · OpenStreetMap"
    s.font      = Font(name="Arial", italic=True, color="676B6C", size=9)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 16

    headers   = ["#", "Business Name", "Type", "Address", "Phone", "Opening Hours", "Website?", "Google Maps"]
    colWidths = [4, 38, 16, 42, 18, 28, 12, 40]

    for col, (h, w) in enumerate(zip(headers, colWidths), 1):
        cell           = ws1.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, color="212023", size=10)
        cell.fill      = accentFill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thinBorder
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[3].height = 22

    for i, p in enumerate(noWebsite, 1):
        row  = i + 3
        fill = altFill if i % 2 == 0 else None
        data = [
            i, p["name"], p["type"], p["address"] or "—",
            p["phone"] or "—", p["hours"] or "—", "No Website", p["mapsUrl"]
        ]
        for col, val in enumerate(data, 1):
            cell           = ws1.cell(row=row, column=col, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.border    = thinBorder
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [2, 3, 4, 6]))
            if fill:
                cell.fill = fill
            if col == 8 and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.value     = "Open in Maps"
                cell.font      = Font(name="Arial", size=10, color="0563C1", underline="single")
        ws1.row_dimensions[row].height = 20
    ws1.freeze_panes = "A4"

    ws2        = wb.create_sheet("All Places")
    ws2.merge_cells("A1:I1")
    t2          = ws2["A1"]
    t2.value    = f"All Places — {city} · {catLabel}  |  Total: {len(allPlaces)}  |  No Website: {len(noWebsite)}"
    t2.font     = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    t2.fill     = darkFill
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    headers2   = ["#", "Business Name", "Type", "Address", "Phone", "Hours", "Has Website?", "Website URL", "Maps"]
    colWidths2 = [4, 35, 15, 40, 18, 26, 13, 35, 20]

    for col, (h, w) in enumerate(zip(headers2, colWidths2), 1):
        cell           = ws2.cell(row=2, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, color="212023", size=10)
        cell.fill      = accentFill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thinBorder
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 20

    greenFill = PatternFill("solid", start_color="D4EDDA")
    redFill   = PatternFill("solid", start_color="F8D7DA")

    for i, p in enumerate(allPlaces, 1):
        row  = i + 2
        data = [
            i, p["name"], p["type"], p["address"] or "—",
            p["phone"] or "—", p["hours"] or "—",
            "Yes" if p["hasWebsite"] else "No",
            p["website"] or "—", p["mapsUrl"]
        ]
        for col, val in enumerate(data, 1):
            cell           = ws2.cell(row=row, column=col, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.border    = thinBorder
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [2, 4, 8]))
            if col == 7:
                cell.fill = greenFill if p["hasWebsite"] else redFill
            if col == 9 and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.value     = "Open"
                cell.font      = Font(name="Arial", size=9, color="0563C1", underline="single")
            if col == 8 and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font      = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws2.row_dimensions[row].height = 18

    ws2.freeze_panes = "A3"
    return wb